from typing import ClassVar, Final

from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, StrictStr

from .core.styling_components import TB, Backgrounds, Borders, Comment, Link, T
from .core.xlsx_base import BaseXLSXCellData, BaseXLSXDocument, BaseXLSXSheet
from .utils import column_generator, ensure_correct_instance

_NUMBER_OF_COLUMNS_TO_PREFILL: Final[int] = 5


class DatasetDescriptionParams(BaseModel):
    name: StrictStr = Field(
        ...,
        description=(
            "Descriptive title for the data set. Equivalent to the title of a scientific "
            "paper. The metadata associated with the published version of this dataset "
            "does not currently make use of this field."
        ),
    )
    description: StrictStr = Field(
        ...,
        description=(
            "NOTE This field is not currently used when publishing a SPARC dataset. "
            "Brief description of the study and the data set. Equivalent to the "
            "abstract of a scientific paper. Include the rationale for the approach, "
            "the types of data collected, the techniques used, formats and number of "
            "files and an approximate size. The metadata associated with the published "
            "version of this dataset does not currently make use of this field."
        ),
    )


class SheetFirstDatasetDescriptionV2(BaseXLSXSheet):
    name = "Sheet1"
    cell_styles: ClassVar[list[tuple[str, BaseXLSXCellData]]] = []

    def assemble_data_for_template(
        self, template_data: BaseModel
    ) -> list[tuple[str, BaseXLSXCellData]]:
        static_cells: list[tuple[str, BaseXLSXCellData]] = [
            (
                "A1",
                TB("Metadata element")
                | Comment(
                    "This metadata will be made available to the public as soon as the data set it submitted.",
                    "M Martone",
                    width=400,
                ),
            ),
            ("B1", TB("Description")),
            ("C1", TB("Example")),
            ("D1", TB("Value")),
            ("A2", TB("Metadata Version")),
            ("B2", T("2.1.0")),
            ("C2", T("2.1.0")),
            ("D2", T("2.1.0")),
            ("A3", TB("Type")),
            (
                "B3",
                T(
                    "The type of this dataset, specifically whether it is experimental or computation. "
                    "The only valid values are experimental or computational. If experimental subjects are "
                    "required, if computational, subjects are not required. Set to experimental by default, "
                    "if you are submitting a computational study be sure to change it."
                ),
            ),
            ("C3", T("experimental")),
            ("D3", T("computational")),
            ("A4", TB("Basic information")),
            ("A5", TB("    Title")),
            (
                "B5",
                T(
                    "Descriptive title for the data set. Equivalent to the title of a scientific paper. "
                    "The metadata associated with the published version of this dataset does not currently "
                    "make use of this field."
                ),
            ),
            ("C5", T("My SPARC dataset")),
            ("A6", TB("    Subtitle")),
            (
                "B6",
                T(
                    "    NOTE This field is not currently used when publishing a SPARC dataset. Brief description"
                    " of the study and the data set. Equivalent to the abstract of a scientific paper. Include "
                    "the rationale for the approach, the types of data collected, the techniques used, "
                    "formats and number of files and an approximate size. The metadata associated with "
                    "the published version of this dataset does not currently make use of this field."
                ),
            ),
            (
                "C6",
                T("A really cool dataset that I collected to answer some question."),
            ),
            ("A7", TB("    Keywords")),
            ("B7", T("A set of keywords to assist in search.")),
            ("C7", T("spinal cord, electrophysiology, RNA-seq, mouse")),
            ("A8", TB("    Funding")),
            ("B8", T("Funding sources")),
            ("C8", T("OT2OD025349")),
            ("A9", TB("    Acknowledgments")),
            ("B9", T("Acknowledgments beyond funding and contributors")),
            ("C9", T("Thank you everyone!")),
            ("A10", TB("Study information")),
            ("A11", TB("    Study purpose")),
            (
                "B11",
                T("A description of the study purpose for the structured abstract."),
            ),
            (
                "C11",
                T(
                    "This study was conducted to demonstrate data wranglers how to fill out dataset "
                    "templates."
                ),
            ),
            ("A12", TB("    Study data collection")),
            (
                "B12",
                T(
                    "A description of the study data collection process for this dataset. Used to "
                    "generate the structured abstract."
                ),
            ),
            (
                "C12",
                T(
                    "Using an earlier version of this template we measured how much it confused "
                    "data wranglers by counting the number of emails we had to exchange with them "
                    "in order to fill it out."
                ),
            ),
            ("A13", TB("    Study primary conclusion")),
            (
                "B13",
                T(
                    "A description of the primary conclusion drawn from the study for the structured abstract."
                ),
            ),
            (
                "C13",
                T(
                    "The primary conclusion of this study is that it is hard to make a good dataset template."
                ),
            ),
            ("A14", TB("    Study organ system")),
            (
                "B14",
                T(
                    "The major organ systems related to this study: autonomic ganglion, brain, colon, heart, "
                    "intestine, kidney, large intestine, liver, lower urinary tract, lung, nervous system, "
                    "pancreas, peripheral nervous system, small intestine, spinal cord, spleen, stomach, "
                    "sympathetic nervous system, urinary bladder"
                ),
            ),
            ("C14", T("patch clamp")),
            ("A15", TB("    Study approach")),
            ("B15", T("The experimental approach or approaches taken in this study.")),
            ("C15", T("electrophysiology")),
            ("A16", TB("    Study technique")),
            ("B16", T("The experimental techniques used in this study.")),
            ("C16", T("patch clamp")),
            ("A17", TB("    Study collection title")),
            (
                "B17",
                T("Title of the larger collection of to which this dataset belongs."),
            ),
            ("C17", T("My SPARC research study")),
            ("A18", TB("Contributor information")),
            ("A19", TB("    Contributor name")),
            (
                "B19",
                T(
                    "Name of any contributors to the dataset.  These individuals need not have been authors "
                    "on any publications describing the data, but should be acknowledged for their role in "
                    "producing and publishing the data set.  If more than one, add each contributor in a new column."
                ),
            ),
            ("C19", T("Last, First Middle")),
            ("A20", TB("    Contributor ORCiD")),
            (
                "B20",
                Link(
                    "ORCiD ID. If you don't have an ORCiD, we suggest you sign up for one.",
                    "https://orcid.org/",
                ),
            ),
            ("C20", T("https://orcid.org/0000-0002-5497-0243")),
            ("A21", TB("    Contributor affiliation")),
            ("B21", T("Institutional affiliation for contributors")),
            ("C21", T("https://ror.org/0168r3w48")),
            ("A22", TB("    Contributor role")),
            (
                "B22",
                T(
                    "Contributor role. At most one PrincipalInvestigator and at least one CorrespondingAuthor "
                    "are required. These roles are provided by the Data Cite schema. Options are: PrincipalInvestigator, "
                    "Creator, CoInvestigator, CorrespondingAuthor, DataCollector, DataCurator, DataManager, Distributor, "
                    "Editor, Producer, ProjectLeader, ProjectManager, ProjectMember, RelatedPerson, Researcher, "
                    "ResearchGroup, Sponsor, Supervisor, WorkPackageLeader, Other."
                ),
            ),
            ("C22", T("DataCollector")),
            ("A23", TB("Related protocol, paper, dataset, etc.")),
            ("A24", TB("    Identifier description")),
            ("B24", T("A description of the referent of the related identifier.")),
            ("C24", T("The protocol use to generate this dataset.")),
            ("A25", TB("    Relation type")),
            (
                "B25",
                T(
                    "The relationship that this dataset has to the related identifier. For example the originating "
                    "article would be this dataset IsDescribedBy originating article. The SPARC specific list is: "
                    "IsProtocolFor, HasProtocol, IsSoftwareFor, HasSoftware. The DataCite list is: IsCitedBy, Cites, "
                    "IsSupplementTo, IsSupplementedBy, IsContinuedByContinues, IsDescribedBy, Describes, HasMetadata, "
                    "IsMetadataFor, HasVersion, IsVersionOf, IsNewVersionOf, IsPreviousVersionOf, IsPartOf, HasPart, "
                    "IsPublishedIn, IsReferencedBy, References, IsDocumentedBy, Documents, IsCompiledBy, Compiles, "
                    "IsVariantFormOf, IsOriginalFormOf, IsIdenticalTo, IsReviewedBy, Reviews, IsDerivedFrom, IsSourceOf, "
                    "IsRequiredBy, Requires, IsObsoletedBy, Obsoletes."
                ),
            ),
            ("C25", T("HasProtocol")),
            ("A26", TB("    Identifier")),
            ("B26", T("The identifier for something related to this dataset.")),
            ("C26", T("https://doi.org/10.13003/5jchdy")),
            ("A27", TB("    Identifier type")),
            ("B27", T("The type of the identifier.")),
            ("C27", T("DOI")),
            ("A28", TB("    Identifier type")),
            ("A29", TB("    Number of subjects")),
            (
                "B29",
                T(
                    "Number of unique subjects in this dataset, should match subjects metadata file. Only required "
                    "for experimental datasets."
                ),
            ),
            ("C29", T("1")),
            ("A30", TB("    Number of samples")),
            (
                "B30",
                T(
                    "Number of unique samples in this dataset, should match samples metadata file. Set to zero if "
                    "there are no samples. Only required for experimental datasets."
                ),
            ),
            ("C30", T("0")),
        ]

        dataset_description: DatasetDescriptionParams = ensure_correct_instance(
            template_data, DatasetDescriptionParams
        )

        data_cells: list[tuple[str, BaseXLSXCellData]] = [
            ("D5", T(dataset_description.name)),
            ("D6", T(dataset_description.description)),
        ]

        style_cells: list[tuple[str, BaseXLSXCellData]] = [
            ("A1:D2", Backgrounds.blue | Borders.light_grid),
            ("A3:C3", Backgrounds.green_light | Borders.light_grid),
            ("A4:D4", Backgrounds.gray_dark | Borders.light_grid),
            ("A5:C7", Backgrounds.green_light | Borders.light_grid),
            ("A8:C9", Backgrounds.yellow | Borders.light_grid),
            ("A10:D10", Backgrounds.gray_dark | Borders.light_grid),
            ("A11:C16", Backgrounds.green_light | Borders.light_grid),
            ("A17:C17", Backgrounds.yellow | Borders.light_grid),
            ("A18:D18", Backgrounds.gray_dark | Borders.light_grid),
            ("A19:C22", Backgrounds.green_light | Borders.light_grid),
            ("A23:D23", Backgrounds.gray_dark | Borders.light_grid),
            ("A24:C27", Backgrounds.green_light | Borders.light_grid),
            ("A28:D28", Backgrounds.gray_dark | Borders.light_grid),
            ("A29:C30", Backgrounds.green_light | Borders.light_grid),
        ]
        style_cells.extend(
            (f"{c}1", T(f"Value {i}") | Backgrounds.blue | Borders.light_grid)
            for i, c in enumerate(column_generator(5, _NUMBER_OF_COLUMNS_TO_PREFILL), 1)
        )
        empty_background_cells: list[tuple[str, BaseXLSXCellData]] = [
            (
                f"E2:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}3",
                Backgrounds.gray_background,
            ),
            (
                f"E4:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}4",
                Backgrounds.gray_dark,
            ),
            (
                f"E5:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}6",
                Backgrounds.gray_background,
            ),
            (
                f"E9:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}9",
                Backgrounds.gray_background,
            ),
            (
                f"E10:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}10",
                Backgrounds.gray_dark,
            ),
            (
                f"E11:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}13",
                Backgrounds.gray_background,
            ),
            (
                f"E17:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}17",
                Backgrounds.gray_background,
            ),
            (
                f"E18:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}18",
                Backgrounds.gray_dark,
            ),
            (
                f"E23:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}23",
                Backgrounds.gray_dark,
            ),
            (
                f"E28:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}28",
                Backgrounds.gray_dark,
            ),
            (
                f"E29:{get_column_letter(4+_NUMBER_OF_COLUMNS_TO_PREFILL)}30",
                Backgrounds.gray_background,
            ),
        ]

        return static_cells + data_cells + style_cells + empty_background_cells

    column_dimensions: ClassVar[dict[str, int]] = {
        "A": 40,
        "B": 60,
        "C": 40,
        "D": 40,
        **{f"{c}": 40 for c in column_generator(5, _NUMBER_OF_COLUMNS_TO_PREFILL)},
    }


class DatasetDescriptionXLSXDocument(BaseXLSXDocument):
    file_name = "dataset_description.xlsx"
    sheet1 = SheetFirstDatasetDescriptionV2()
