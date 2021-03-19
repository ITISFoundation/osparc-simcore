import inspect

from typing import Dict, Generator, Tuple, Any, Set
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Alignment
from openpyxl.cell import Cell


def _base_value_or(self_var: Any, entry_var: Any) -> Any:
    """
    For some properties it is more convenient
    to also have their fields merged, like:
    borders and alignment
    """
    if isinstance(self_var, Border):
        return Border(
            left=self_var.left or entry_var.left,
            right=self_var.right or entry_var.right,
            top=self_var.top or entry_var.top,
            bottom=self_var.bottom or entry_var.bottom,
            diagonal=self_var.diagonal or entry_var.diagonal,
            diagonal_direction=(
                self_var.diagonal_direction or entry_var.diagonal_direction
            ),
            vertical=self_var.vertical or entry_var.vertical,
            horizontal=self_var.horizontal or entry_var.horizontal,
            diagonalUp=self_var.diagonalUp or entry_var.diagonalUp,
            diagonalDown=self_var.diagonalDown or entry_var.diagonalDown,
            outline=self_var.outline or entry_var.outline,
            start=self_var.start or entry_var.start,
            end=self_var.end or entry_var.end,
        )
    if isinstance(self_var, Alignment):
        return Alignment(
            horizontal=self_var.horizontal or entry_var.horizontal,
            vertical=self_var.vertical or entry_var.vertical,
            textRotation=self_var.textRotation or entry_var.textRotation,
            wrapText=self_var.wrapText or entry_var.wrapText,
            shrinkToFit=self_var.shrinkToFit or entry_var.shrinkToFit,
            indent=self_var.indent or entry_var.indent,
            relativeIndent=self_var.relativeIndent or entry_var.relativeIndent,
            justifyLastLine=self_var.justifyLastLine or entry_var.justifyLastLine,
            readingOrder=self_var.readingOrder or entry_var.readingOrder,
            text_rotation=self_var.text_rotation or entry_var.text_rotation,
            wrap_text=self_var.wrap_text or entry_var.wrap_text,
            shrink_to_fit=self_var.shrink_to_fit or entry_var.shrink_to_fit,
        )

    return self_var or entry_var


class BaseXLSXCellData:
    """
    It is used to "or" `openpyxl.cell.Cell` properties where
    the below conditions apply. Refer to `openpyxl.cell.Cell` for
    supported fields.

    The idea is to have cells which are additive so properties
    are chained via "|" operator.
    Values are defined also via class level properties so they
    can easily be overwritten, via class inheritance and constructors.

    All properties will be merged with "or" rule, the below is True:

        BaseXLSXCellData(value="text") | BaseXLSXCellData(font=Font())
            == BaseXLSXCellData(value="text", font=Font())

    The "or" operator has left-to-right associativity, the below is True:

        BaseXLSXCellData(value="text") | BaseXLSXCellData(value="other_text")
            == BaseXLSXCellData(value="text")
    """

    def __init__(self, **kwargs):
        # gather all class attributes
        # a class attribute is considered a member not starting with "__" or "_"
        for name, value in vars(self.__class__).items():
            if name.startswith("__") or name.startswith("_"):
                continue
            self.__setattr__(name, value)

        # do not overwrite class level definition if already exists
        for name, value in kwargs.items():
            to_store = self.__getattribute__(name) if value is None else value
            self.__setattr__(name, to_store)

        print(">>>>setup", self.__class__.__name__, vars(self).keys())

    def __or__(self, other):
        """
        Applies or operation to all shared attributes and
        copy the unshared from both.
        """
        self_vars = vars(self)
        other_vars = vars(other)

        self_vars_keys = set(self_vars.keys())
        other_vars_keys = set(other_vars.keys())

        keys_in_common = self_vars_keys & other_vars_keys
        keys_self_vars_to_copy = self_vars_keys - keys_in_common
        keys_other_vars_to_copy = other_vars_keys - keys_in_common

        # apply or to common properties
        merged_values = {
            k: _base_value_or(self_vars[k], other_vars[k]) for k in keys_in_common
        }

        # copy properties which are not shared
        for key in keys_self_vars_to_copy:
            merged_values[key] = self_vars[key]
        for key in keys_other_vars_to_copy:
            merged_values[key] = other_vars[key]

        return BaseXLSXCellData(**merged_values)

    def __repr__(self):
        """Only outputs not None attributes"""
        formatted_vars = ", ".join(
            [f"{x[0]}={x[1]}" for x in vars(self).items() if x[1] is not None]
        )
        return f"<{self.__class__.__name__} {formatted_vars}>"


class BaseXLSXSheet:
    # name of the sheet
    name: str = None
    # cell style contents
    cell_styles: Dict[str, Dict[str, BaseXLSXCellData]] = None
    # cell merge contents
    cell_merge: Set[str] = set()
    # column length
    column_dimensions: Dict[str, int] = {}

    def _check_attribute(self, attribute_name: str):
        if getattr(self, attribute_name) is None:
            raise ValueError(f"'{attribute_name}' attribute is None, please define it")

    def __init__(self):
        self._check_attribute("cell_styles")
        self._check_attribute("name")

    def __repr__(self):
        return f"<{self.__class__.__name__} name={self.name}, cell_styles={self.cell_styles}"


def _update_cell(cell: Cell, data: BaseXLSXCellData) -> None:
    """Extract properties from the cell_styles and apply them to the cell"""
    for name, value in vars(data).items():
        cell.__setattr__(name, value)


def _update_entry_in_cell(
    target: Dict[str, BaseXLSXCellData],
    address: str,
    new_entry: BaseXLSXCellData,
) -> None:
    """
    There may be multiple entires for the same cell, coming from different sources.
    It is useful for applying styling to existing cells and storing values
    """
    exiting_entry = target.get(address, None)
    print(">>>>>", type(exiting_entry), type(new_entry), "||||||||")
    target[address] = (
        new_entry if exiting_entry is None else (exiting_entry | new_entry)
    )


def _assemble_workbook(
    sheets_entries: Generator[Tuple[str, Any], None, None]
) -> Workbook:
    workbook = Workbook()

    for _, sheet_data in sheets_entries:
        sheet_data: BaseXLSXSheet = sheet_data
        sheet_name = sheet_data.name

        xls_sheet = workbook.create_sheet(sheet_name)

        single_cells_cell_styles: Dict[str, BaseXLSXCellData] = {}

        # flatten out ranges like A1:B20 in single entries
        for cell_address, entry in sheet_data.cell_styles.items():
            if ":" in cell_address:
                for cell_row in xls_sheet[cell_address]:
                    for cell in cell_row:
                        _update_entry_in_cell(
                            target=single_cells_cell_styles,
                            address=cell.coordinate,
                            new_entry=entry,
                        )
            else:
                _update_entry_in_cell(
                    target=single_cells_cell_styles,
                    address=cell_address,
                    new_entry=entry,
                )

        # finally apply data from cell cell_styles to xls cells
        for cell_address, entry in single_cells_cell_styles.items():
            _update_cell(xls_sheet[cell_address], entry)

        # apply column widths
        for column, width in sheet_data.column_dimensions.items():
            xls_sheet.column_dimensions[column].width = width

        # apply cell merging
        for to_merge in sheet_data.cell_merge:
            xls_sheet.merge_cells(to_merge)

    # remove the default sheet
    sheet_to_remove = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
    workbook.remove(sheet_to_remove)

    return workbook


class BaseXLSXDocument:
    def __init__(self, *args):
        for k, entry in enumerate(args):
            self.__dict__[f"__sheet__entry__{k}"] = entry

    def _get_sheets(self) -> Generator[Tuple[str, Any], None, None]:
        for member in inspect.getmembers(self):
            if isinstance(member[1], BaseXLSXSheet):
                yield member

    def __repr__(self):
        formatted_sheets = "\n\t".join([f"{x[0]}={x[1]}" for x in self._get_sheets()])
        return f"<{self.__class__.__name__}\n\t{formatted_sheets}>"

    def _generate_document(self) -> Workbook:
        return _assemble_workbook(self._get_sheets())

    def save_document(self, file_path: Path) -> None:
        workbook = self._generate_document()
        workbook.save(file_path)
